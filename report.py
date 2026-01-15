#!/usr/bin/env python3
# report.py
# Generate publishable Safety Hazard Report (XLSX + HTML) from hazards_enriched*.json
#
# Features:
# - Stable XLSX export with multiple sheets:
#   Hazards / Requirements / Verification / Traceability / Summary / SystemInput / TopRisks
# - Review workflow columns:
#   review_status / reviewer / review_notes / decision_date
# - Risk scoring (RPN) and sorting by risk by default
# - Executive summary (Top N risks) in HTML and in TopRisks sheet

from __future__ import annotations

import argparse
import json
import math
import re
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Tuple

try:
    import pandas as pd  # type: ignore
except Exception:  # pragma: no cover
    pd = None  # type: ignore

try:
    import yaml  # type: ignore
except Exception:  # pragma: no cover
    yaml = None  # type: ignore

TRUNC_MARKERS = ("...", "…")

# Simple weights (adjust later to match a customer standard)
SEVERITY_W = {"Low": 1, "Medium": 2, "High": 3, "Critical": 4}
LIKELIHOOD_W = {"Low": 1, "Medium": 2, "High": 3}
DETECT_W = {"High": 1, "Medium": 2, "Low": 3}

DEFAULT_REVIEW_STATUS = "Draft"  # Draft/Reviewed/Accepted/Rejected


def load_json(path: Path) -> Dict[str, Any]:
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)


def ensure_dir(p: Path) -> None:
    p.mkdir(parents=True, exist_ok=True)


def norm_text(s: Any) -> str:
    if s is None:
        return ""
    if isinstance(s, (int, float, bool)):
        return str(s)
    return str(s).strip()


def join_lines(x: Any) -> str:
    """Join list[str] into newline-separated bullets; pass through strings."""
    if x is None:
        return ""
    if isinstance(x, list):
        items = [norm_text(i) for i in x if norm_text(i)]
        return "\n".join(f"- {i}" for i in items)
    return norm_text(x)


def tokenize(s: str) -> List[str]:
    return re.findall(r"[a-z0-9]+", s.lower())


def token_set_similarity(a: str, b: str) -> float:
    ta = set(tokenize(a))
    tb = set(tokenize(b))
    if not ta or not tb:
        return 0.0
    inter = ta.intersection(tb)
    union = ta.union(tb)
    return len(inter) / max(1, len(union))


def system_input_raw_yaml(doc: Dict[str, Any]) -> str:
    ydoc = doc.get("system_input_yaml")
    if isinstance(ydoc, dict):
        if yaml is not None:
            return yaml.safe_dump(ydoc, sort_keys=False, allow_unicode=True)
        return json.dumps(ydoc, ensure_ascii=False, indent=2)
    return norm_text(doc.get("system_input"))


def _join_listish(v: Any) -> str:
    if isinstance(v, list):
        items = [norm_text(x) for x in v if norm_text(x)]
        return ", ".join(items)
    return norm_text(v)


def odd_summary_lines(doc: Dict[str, Any]) -> List[str]:
    ydoc = doc.get("system_input_yaml")
    if not isinstance(ydoc, dict):
        return [
            "ODD: unknown",
            "Actors: unknown",
            "Autonomy: unknown",
            "Constraints: unknown",
        ]

    op_ctx = ydoc.get("operational_context", {}) if isinstance(ydoc.get("operational_context"), dict) else {}
    odd = op_ctx.get("operational_design_domain", {}) if isinstance(op_ctx.get("operational_design_domain"), dict) else {}

    site = _join_listish(odd.get("site_type"))
    ground = _join_listish(odd.get("ground_conditions"))
    weather = _join_listish(odd.get("weather_lighting"))
    traffic = _join_listish(odd.get("traffic_and_people"))

    odd_parts = []
    if site:
        odd_parts.append(f"site={site}")
    if ground:
        odd_parts.append(f"ground={ground}")
    if weather:
        odd_parts.append(f"weather={weather}")
    if traffic:
        odd_parts.append(f"traffic={traffic}")
    odd_line = "ODD: " + ("; ".join(odd_parts) if odd_parts else "unknown")

    actors = ydoc.get("actors", {}) if isinstance(ydoc.get("actors"), dict) else {}
    humans = actors.get("humans", []) if isinstance(actors.get("humans"), list) else []
    ext = actors.get("external_systems", []) if isinstance(actors.get("external_systems"), list) else []
    actor_names = []
    for h in humans:
        if isinstance(h, dict):
            role = norm_text(h.get("role"))
            if role:
                actor_names.append(role)
    for s in ext:
        if isinstance(s, dict):
            name = norm_text(s.get("name"))
            if name:
                actor_names.append(name)
    actors_line = "Actors: " + (", ".join(actor_names) if actor_names else "unknown")

    sys_block = ydoc.get("system", {}) if isinstance(ydoc.get("system"), dict) else {}
    autonomy = norm_text(sys_block.get("autonomy_level"))
    if not autonomy:
        modes = ydoc.get("modes_and_states", {})
        if isinstance(modes, dict):
            autonomy = _join_listish(modes.get("modes"))
    autonomy_line = "Autonomy: " + (autonomy if autonomy else "unknown")

    constraints = op_ctx.get("constraints", None)
    if constraints is None:
        constraints = sys_block.get("constraints", None)
    constraints_line = "Constraints: " + (_join_listish(constraints) if constraints is not None else "unknown")

    return [odd_line, actors_line, autonomy_line, constraints_line]


def odd_tags(doc: Dict[str, Any]) -> List[str]:
    ydoc = doc.get("system_input_yaml")
    if not isinstance(ydoc, dict):
        return []
    op_ctx = ydoc.get("operational_context", {}) if isinstance(ydoc.get("operational_context"), dict) else {}
    odd = op_ctx.get("operational_design_domain", {}) if isinstance(op_ctx.get("operational_design_domain"), dict) else {}
    tags = []
    for key in ["site_type", "ground_conditions", "weather_lighting", "traffic_and_people"]:
        v = odd.get(key)
        if isinstance(v, list):
            for item in v:
                item_s = norm_text(item)
                if item_s:
                    tags.append(f"{key}:{item_s}")
        elif norm_text(v):
            tags.append(f"{key}:{norm_text(v)}")
    return tags


def extract_assumptions(doc: Dict[str, Any]) -> List[Dict[str, str]]:
    ydoc = doc.get("system_input_yaml")
    if not isinstance(ydoc, dict):
        return []
    assumptions = ydoc.get("assumptions", {}) if isinstance(ydoc.get("assumptions"), dict) else {}
    items = assumptions.get("list", []) if isinstance(assumptions.get("list"), list) else []
    out = []
    for it in items:
        if not isinstance(it, dict):
            continue
        aid = norm_text(it.get("id"))
        statement = norm_text(it.get("statement"))
        if aid and statement:
            out.append({"id": aid, "statement": statement})
    return out


def assumption_ids_for(hazard_text: str, assumptions: List[Dict[str, str]]) -> List[str]:
    ids = []
    htoks = set(tokenize(hazard_text))
    for a in assumptions:
        atoks = set(tokenize(a.get("statement", "")))
        if not atoks:
            continue
        overlap = len(htoks.intersection(atoks))
        if overlap >= 2:
            ids.append(a.get("id", ""))
    return [i for i in ids if i]


def has_trunc(s: str) -> bool:
    s2 = s.strip()
    if not s2:
        return False
    if any(s2.endswith(m) for m in TRUNC_MARKERS):
        return True
    if "..." in s2:
        return True
    return False


def make_ids(hazard_id: str) -> Tuple[str, str]:
    sr = f"SR-{hazard_id}-01"
    vr = f"VR-{hazard_id}-01"
    return sr, vr


def weight(map_: Dict[str, int], label: str, default: int) -> int:
    return map_.get(label, default)


def risk_score(severity: str, likelihood: str, detectability: str) -> int:
    s = weight(SEVERITY_W, severity, 2)
    l = weight(LIKELIHOOD_W, likelihood, 2)
    d = weight(DETECT_W, detectability, 2)
    return int(s * l * d)


def risk_band(score: int) -> str:
    # Interpretable bands; tweak later
    if score >= 30:
        return "Extreme"
    if score >= 18:
        return "High"
    if score >= 8:
        return "Medium"
    return "Low"


def to_rows(doc: Dict[str, Any]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    hazards = doc.get("hazards", [])
    if not isinstance(hazards, list):
        hazards = []
    assumptions = extract_assumptions(doc)
    odd_tag_list = odd_tags(doc)

    for h in hazards:
        hazard_text = " ".join(
            [
                norm_text(h.get("trigger_condition")),
                norm_text(h.get("primary_failure")),
                norm_text(h.get("final_impact")),
            ]
        )
        assumption_ids = assumption_ids_for(hazard_text, assumptions)
        odd_tags_joined = ", ".join(odd_tag_list) if odd_tag_list else ""

        hid = norm_text(h.get("hazard_id"))
        sr_id, vr_id = make_ids(hid)

        sev = norm_text(h.get("severity"))
        lik = norm_text(h.get("likelihood"))
        det = norm_text(h.get("detectability"))
        rpn = risk_score(sev, lik, det)

        obs = join_lines(h.get("observables"))
        mit = join_lines(h.get("mitigations"))
        req = norm_text(h.get("safety_requirement"))
        vr_idea = norm_text(h.get("verification_idea"))
        measures = join_lines(h.get("observables"))
        warnings = []
        if not req:
            warnings.append("Missing safety_requirement")
        if not vr_idea:
            warnings.append("Missing verification_idea")
        warning_text = "; ".join(warnings)

        # Review workflow fields (optional)
        review = h.get("review", {}) if isinstance(h.get("review"), dict) else {}

        rows.append(
            {
                "hazard_id": hid,
                "category": norm_text(h.get("category")),
                "subtype": norm_text(h.get("subtype")),
                "task_phase": norm_text(h.get("task_phase")),
                "severity": sev,
                "likelihood": lik,
                "detectability": det,
                "risk_rpn": rpn,
                "risk_band": risk_band(rpn),
                "affected_components": ", ".join([norm_text(x) for x in (h.get("affected_components") or [])]),
                "trigger_condition": norm_text(h.get("trigger_condition")),
                "primary_failure": norm_text(h.get("primary_failure")),
                "propagation_chain": join_lines(h.get("propagation_chain")),
                "final_impact": norm_text(h.get("final_impact")),
                "safety_requirement_id": norm_text(h.get("safety_requirement_id")) or sr_id,
                "verification_id": norm_text(h.get("verification_id")) or vr_id,
                "observables": obs,
                "measures": measures,
                "mitigations": mit,
                "safety_requirement": req,
                "verification_idea": vr_idea,
                "flag_req_truncated": "YES" if has_trunc(req) else "",
                "flag_verif_truncated": "YES" if has_trunc(vr_idea) else "",
                "assumption_ids": ", ".join(assumption_ids),
                "odd_tags": odd_tags_joined,
                "review_warnings": warning_text,
                "review_status": norm_text(review.get("status")) or DEFAULT_REVIEW_STATUS,
                "reviewer": norm_text(review.get("reviewer")),
                "review_notes": norm_text(review.get("notes")),
                "decision_date": norm_text(review.get("decision_date")),
            }
        )

    # Default sort: highest risk first, then severity, then phase, then id
    rows.sort(
        key=lambda r: (
            -int(r.get("risk_rpn", 0) or 0),
            -SEVERITY_W.get(r.get("severity", ""), 0),
            r.get("task_phase", ""),
            r.get("hazard_id", ""),
        )
    )
    return rows


def to_requirements_rows(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    return [
        {
            "safety_requirement_id": r.get("safety_requirement_id", ""),
            "hazard_id": r.get("hazard_id", ""),
            "category": r.get("category", ""),
            "task_phase": r.get("task_phase", ""),
            "severity": r.get("severity", ""),
            "risk_rpn": r.get("risk_rpn", ""),
            "risk_band": r.get("risk_band", ""),
            "affected_components": r.get("affected_components", ""),
            "safety_requirement": r.get("safety_requirement", ""),
            "review_status": r.get("review_status", ""),
            "reviewer": r.get("reviewer", ""),
            "decision_date": r.get("decision_date", ""),
            "flag_req_truncated": r.get("flag_req_truncated", ""),
        }
        for r in rows
    ]


def to_verification_rows(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    return [
        {
            "verification_id": r.get("verification_id", ""),
            "hazard_id": r.get("hazard_id", ""),
            "category": r.get("category", ""),
            "task_phase": r.get("task_phase", ""),
            "severity": r.get("severity", ""),
            "risk_rpn": r.get("risk_rpn", ""),
            "risk_band": r.get("risk_band", ""),
            "observables": r.get("observables", ""),
            "mitigations": r.get("mitigations", ""),
            "verification_idea": r.get("verification_idea", ""),
            "review_status": r.get("review_status", ""),
            "reviewer": r.get("reviewer", ""),
            "decision_date": r.get("decision_date", ""),
            "flag_verif_truncated": r.get("flag_verif_truncated", ""),
        }
        for r in rows
    ]


def to_traceability_rows(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    return [
        {
            "hazard_id": r.get("hazard_id", ""),
            "safety_requirement_id": r.get("safety_requirement_id", ""),
            "verification_id": r.get("verification_id", ""),
            "severity": r.get("severity", ""),
            "risk_rpn": r.get("risk_rpn", ""),
            "risk_band": r.get("risk_band", ""),
            "task_phase": r.get("task_phase", ""),
            "category": r.get("category", ""),
            "assumption_ids": r.get("assumption_ids", ""),
            "odd_tags": r.get("odd_tags", ""),
            "review_status": r.get("review_status", ""),
        }
        for r in rows
    ]


def summarize(rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    cat = Counter(r["category"] for r in rows)
    sev = Counter(r["severity"] for r in rows)
    phase = Counter(r["task_phase"] for r in rows)
    band = Counter(r["risk_band"] for r in rows)

    trunc_req = sum(1 for r in rows if r.get("flag_req_truncated") == "YES")
    trunc_ver = sum(1 for r in rows if r.get("flag_verif_truncated") == "YES")

    return {
        "total_hazards": len(rows),
        "category_distribution": dict(cat),
        "severity_distribution": dict(sev),
        "task_phase_top15": dict(phase.most_common(15)),
        "risk_band_distribution": dict(band),
        "truncation_flags": {"safety_requirement_truncated": trunc_req, "verification_idea_truncated": trunc_ver},
    }


def top_risks(rows: List[Dict[str, Any]], n: int) -> List[Dict[str, Any]]:
    return rows[: max(0, n)]


def top_risk_reason(row: Dict[str, Any]) -> str:
    sev = norm_text(row.get("severity"))
    lik = norm_text(row.get("likelihood"))
    det = norm_text(row.get("detectability"))
    if sev in {"Critical", "High"}:
        return "High severity: potential for serious harm or major damage."
    if det == "Low":
        return "Low detectability: failure may be missed until too late."
    if lik == "High":
        return "High likelihood: expected to occur more often."
    return "High combined risk score compared to other hazards."


def recommended_action(row: Dict[str, Any]) -> Tuple[str, str]:
    sev = norm_text(row.get("severity"))
    lik = norm_text(row.get("likelihood"))
    det = norm_text(row.get("detectability"))
    if det == "Low":
        return "Test", "Low detectability: prioritize detection/validation coverage."
    if sev in {"Critical", "High"}:
        return "Design", "High severity: reduce inherent risk at the source."
    if lik == "High":
        return "Process", "High likelihood: improve operational controls and training."
    return "Design", "Overall risk score is high relative to peers."


def hazard_summary_line(row: Dict[str, Any]) -> str:
    trig = norm_text(row.get("trigger_condition"))
    impact = norm_text(row.get("final_impact"))
    if trig and impact:
        return f"{trig} -> {impact}"
    return trig or impact or "Risk scenario requires review."


def distribution_sanity_check(dist: Dict[str, int]) -> Dict[str, Any]:
    counts = [int(v) for v in dist.values() if isinstance(v, int)]
    if not counts:
        return {"max": 0, "min": 0, "max_min_ratio": None, "gini": None, "warning": "No data"}
    max_v = max(counts)
    min_v = min(counts)
    ratio = (max_v / min_v) if min_v > 0 else None
    sorted_counts = sorted(counts)
    n = len(sorted_counts)
    if n == 0:
        gini = None
    else:
        cum = 0.0
        for i, x in enumerate(sorted_counts, start=1):
            cum += i * x
        total = sum(sorted_counts)
        gini = (2 * cum) / (n * total) - (n + 1) / n if total > 0 else None
    warning = ""
    if min_v == max_v and len(counts) > 1:
        warning = "Warning: category counts are perfectly uniform; verify generator balance."
    return {"max": max_v, "min": min_v, "max_min_ratio": ratio, "gini": gini, "warning": warning}


def cluster_hazards(rows: List[Dict[str, Any]], threshold: float = 0.45) -> Dict[str, Dict[str, Any]]:
    clusters: Dict[str, Dict[str, Any]] = {}
    cluster_id = 0
    for r in rows:
        text = " ".join(
            [
                norm_text(r.get("trigger_condition")),
                norm_text(r.get("primary_failure")),
                norm_text(r.get("final_impact")),
            ]
        )
        assigned = None
        for cid, c in clusters.items():
            sim = token_set_similarity(text, c["rep_text"])
            if sim >= threshold:
                assigned = cid
                break
        if assigned is None:
            cluster_id += 1
            assigned = f"C{cluster_id:02d}"
            clusters[assigned] = {"rep_text": text, "items": []}
        clusters[assigned]["items"].append(r)
    return clusters


def apply_cluster_labels(rows: List[Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
    clusters = cluster_hazards(rows)
    for cid, c in clusters.items():
        size = len(c["items"])
        for r in c["items"]:
            r["cluster_id"] = cid
            r["cluster_size"] = size
    return clusters


def apply_action_fields(rows: List[Dict[str, Any]]) -> None:
    for r in rows:
        action, rationale = recommended_action(r)
        r["recommended_action"] = action
        r["action_rationale"] = rationale

def write_html(out_path: Path, doc: Dict[str, Any], rows: List[Dict[str, Any]], summary: Dict[str, Any], topn: int, generated_at: str, clusters: Dict[str, Dict[str, Any]]) -> None:
    system_input = system_input_raw_yaml(doc)
    now = generated_at

    def esc(s: str) -> str:
        return (
            s.replace("&", "&amp;")
            .replace("<", "&lt;")
            .replace(">", "&gt;")
            .replace('"', "&quot;")
            .replace("'", "&#039;")
        )

    def dict_to_ul(d: Dict[str, Any]) -> str:
        items = []
        for k, v in d.items():
            items.append(f"<li><b>{esc(str(k))}</b>: {esc(str(v))}</li>")
        return "<ul>" + "".join(items) + "</ul>"

    def table_html(title: str, anchor: str, cols: List[str], data_rows: List[Dict[str, Any]]) -> str:
        trs = []
        for r in data_rows:
            tds = []
            for c in cols:
                val = norm_text(r.get(c))
                tds.append(f"<td><pre>{esc(val)}</pre></td>")
            trs.append("<tr>" + "".join(tds) + "</tr>")
        return f"""
<div class="card" id="{anchor}">
  <h2>{esc(title)}</h2>
  <table>
    <thead>
      <tr>{''.join([f'<th><pre>{esc(c)}</pre></th>' for c in cols])}</tr>
    </thead>
    <tbody>
      {''.join(trs)}
    </tbody>
  </table>
</div>
""".strip()

    toc = """
<div class="card" id="toc">
  <h2>Contents</h2>
  <ul>
    <li><a href="#summary">Summary</a></li>
    <li><a href="#toprisks">Top risks</a></li>
    <li><a href="#hazards">Hazards</a></li>
    <li><a href="#requirements">Requirements</a></li>
    <li><a href="#verification">Verification</a></li>
    <li><a href="#traceability">Traceability</a></li>
  </ul>
</div>
""".strip()

    empty_note = ""
    if summary["total_hazards"] == 0:
        empty_note = "<div><b>Status</b>: No hazards generated</div>"

    odd_lines = odd_summary_lines(doc)
    odd_block = "<ul>" + "".join(f"<li>{esc(line)}</li>" for line in odd_lines) + "</ul>"

    sanity = distribution_sanity_check(summary["category_distribution"])
    sanity_lines = [
        f"max={sanity['max']}",
        f"min={sanity['min']}",
        f"max/min={sanity['max_min_ratio']}",
        f"gini={sanity['gini']}",
    ]
    sanity_warning = norm_text(sanity.get("warning"))
    sanity_block = "<ul>" + "".join(f"<li>{esc(s)}</li>" for s in sanity_lines) + "</ul>"
    if sanity_warning:
        sanity_block += f"<div class=\"warn\">{esc(sanity_warning)}</div>"

    top5 = rows[:5]
    top5_items = []
    for r in top5:
        top5_items.append(
            f"<li><b>{esc(norm_text(r.get('hazard_id')))}</b>: {esc(hazard_summary_line(r))}"
            f"<br/><b>Action</b>: {esc(norm_text(r.get('recommended_action')))}"
            f" | <b>Why</b>: {esc(norm_text(r.get('action_rationale')))}"
            f" | <b>Assumptions</b>: {esc(norm_text(r.get('assumption_ids')) or 'unknown')}"
            f" | <b>ODD</b>: {esc(norm_text(r.get('odd_tags')) or 'unknown')}"
            f"</li>"
        )
    top5_block = "<ol>" + "".join(top5_items) + "</ol>" if top5_items else "<div>none</div>"

    cluster_items = []
    for cid, c in sorted(clusters.items(), key=lambda x: len(x[1]["items"]), reverse=True)[:5]:
        rep = c["items"][0] if c["items"] else {}
        cluster_items.append(
            f"<li><b>{esc(cid)}</b> ({len(c['items'])} items): "
            f"rep={esc(norm_text(rep.get('hazard_id')))}</li>"
        )
    cluster_block = "<ul>" + "".join(cluster_items) + "</ul>" if cluster_items else "<div>none</div>"

    summary_card = f"""
<div class="card" id="summary">
  <h2>Summary</h2>
  <div><b>Generated</b>: {esc(now)}</div>
  <div><b>ODD summary</b>:</div>
  {odd_block}
  <h3>Decision-ready blocks</h3>
  <div><b>Top 5 risks to act on next</b></div>
  {top5_block}
  <div><b>Top clusters</b></div>
  {cluster_block}
  <div><b>Total hazards</b>: {summary['total_hazards']}</div>
  {empty_note}
  <h3>Risk bands</h3>
  {dict_to_ul(summary['risk_band_distribution'])}
  <h3>Category distribution</h3>
  {dict_to_ul(summary['category_distribution'])}
  <h3>Severity distribution</h3>
  {dict_to_ul(summary['severity_distribution'])}
  <h3>Top task phases (15)</h3>
  {dict_to_ul(summary['task_phase_top15'])}
  <h3>Truncation flags</h3>
  {dict_to_ul(summary['truncation_flags'])}
  <h3>Distribution sanity check</h3>
  {sanity_block}
</div>
""".strip()

    system_card = f"""
<div class="card" id="system">
  <h2>System Input</h2>
  <pre>{esc(system_input)}</pre>
</div>
""".strip()

    top_cols = [
        "hazard_id", "risk_rpn", "risk_band", "severity", "likelihood", "detectability",
        "category", "task_phase", "affected_components", "trigger_condition", "final_impact",
        "safety_requirement_id", "verification_id", "review_status", "why_top",
    ]

    hazards_cols = [
        "hazard_id", "risk_rpn", "risk_band",
        "category", "subtype", "task_phase",
        "severity", "likelihood", "detectability",
        "affected_components",
        "trigger_condition", "primary_failure", "final_impact",
        "safety_requirement_id", "safety_requirement",
        "verification_id", "verification_idea",
        "observables", "measures", "mitigations",
        "assumption_ids", "odd_tags", "cluster_id", "cluster_size",
        "recommended_action", "action_rationale",
        "review_status", "reviewer", "decision_date", "review_notes",
        "flag_req_truncated", "flag_verif_truncated",
        "review_warnings",
    ]

    req_cols = [
        "safety_requirement_id", "hazard_id",
        "risk_rpn", "risk_band",
        "category", "task_phase", "severity",
        "affected_components",
        "safety_requirement",
        "review_status", "reviewer", "decision_date",
        "flag_req_truncated",
    ]

    ver_cols = [
        "verification_id", "hazard_id",
        "risk_rpn", "risk_band",
        "category", "task_phase", "severity",
        "observables", "mitigations", "verification_idea",
        "review_status", "reviewer", "decision_date",
        "flag_verif_truncated",
    ]

    tr_cols = [
        "hazard_id", "safety_requirement_id", "verification_id",
        "risk_rpn", "risk_band",
        "severity", "task_phase", "category",
        "assumption_ids", "odd_tags",
        "review_status",
    ]

    css = """
<style>
body { font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, Arial, sans-serif; margin: 24px; }
.card { border: 1px solid #ddd; border-radius: 10px; padding: 16px; margin: 16px 0; }
details.card summary { cursor: pointer; }
.warn { color: #8a3b12; font-weight: 600; }
pre { white-space: pre-wrap; word-wrap: break-word; margin: 0; font-family: ui-monospace, SFMono-Regular, Menlo, Monaco, Consolas, "Liberation Mono", "Courier New", monospace; font-size: 12px; }
table { border-collapse: collapse; width: 100%; }
th, td { border: 1px solid #ddd; vertical-align: top; }
th { background: #f6f6f6; position: sticky; top: 0; z-index: 1; }
th pre { font-size: 12px; }
td pre { padding: 8px; }
a { color: #0b57d0; }
.badge { display: inline-block; padding: 2px 8px; border-radius: 999px; border: 1px solid #ccc; margin-left: 8px; font-size: 12px; }
</style>
""".strip()

    top_rows = [dict(r, why_top=top_risk_reason(r)) for r in top_risks(rows, topn)]

    cluster_sections = []
    for cid, c in sorted(clusters.items(), key=lambda x: len(x[1]["items"]), reverse=True):
        rep = c["items"][0] if c["items"] else {}
        title = f"{cid} ({len(c['items'])} items) - rep {norm_text(rep.get('hazard_id'))}"
        cluster_sections.append(
            f"""
<details class="card" id="cluster-{esc(cid)}">
  <summary><b>{esc(title)}</b></summary>
  {table_html(f"Cluster {cid}", f"cluster-{cid}-table", hazards_cols, c["items"])}
</details>
""".strip()
        )

    body = "\n".join(
        [
            toc,
            summary_card,
            system_card,
            table_html(f"Top risks (Top {topn})", "toprisks", top_cols, top_rows),
            "<div class=\"card\" id=\"hazards\"><h2>Hazards (clustered)</h2></div>",
            "\n".join(cluster_sections),
            table_html("Requirements", "requirements", req_cols, to_requirements_rows(rows)),
            table_html("Verification", "verification", ver_cols, to_verification_rows(rows)),
            table_html("Traceability", "traceability", tr_cols, to_traceability_rows(rows)),
        ]
    )

    html = f"""<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8"/>
<meta name="viewport" content="width=device-width, initial-scale=1"/>
<title>Synthetic Danger Report</title>
{css}
</head>
<body>
<h1>Synthetic Danger Report <span class="badge">risk-sorted</span></h1>
{body}
</body>
</html>
"""

    out_path.write_text(html, encoding="utf-8")


def write_xlsx(out_path: Path, rows: List[Dict[str, Any]], summary: Dict[str, Any], doc: Dict[str, Any], topn: int) -> None:
    if pd is None:
        raise RuntimeError("pandas is not available. Install with: pip install pandas openpyxl")

    df = pd.DataFrame(rows)
    req_df = pd.DataFrame(to_requirements_rows(rows))
    ver_df = pd.DataFrame(to_verification_rows(rows))
    tr_df = pd.DataFrame(to_traceability_rows(rows))
    top_rows = [dict(r, why_top=top_risk_reason(r)) for r in top_risks(rows, topn)]
    top_df = pd.DataFrame(top_rows)

    cat_df = pd.DataFrame(list(summary["category_distribution"].items()), columns=["category", "count"]).sort_values(
        "count", ascending=False
    )
    sev_df = pd.DataFrame(list(summary["severity_distribution"].items()), columns=["severity", "count"]).sort_values(
        "count", ascending=False
    )
    band_df = pd.DataFrame(list(summary["risk_band_distribution"].items()), columns=["risk_band", "count"]).sort_values(
        "count", ascending=False
    )
    ph_df = pd.DataFrame(list(summary["task_phase_top15"].items()), columns=["task_phase", "count"])
    trunc_df = pd.DataFrame(list(summary["truncation_flags"].items()), columns=["flag", "count"])

    meta_df = pd.DataFrame(
        [
            ["generated_at", doc.get("_generated_at") or datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["total_hazards", summary["total_hazards"]],
            ["top_n", topn],
            ["sort", "risk_rpn desc"],
        ],
        columns=["key", "value"],
    )
    sys_df = pd.DataFrame([system_input_raw_yaml(doc)], columns=["system_input_raw_yaml"])

    sanity = distribution_sanity_check(summary["category_distribution"])
    sanity_df = pd.DataFrame(
        [
            ["max", sanity.get("max")],
            ["min", sanity.get("min")],
            ["max_min_ratio", sanity.get("max_min_ratio")],
            ["gini", sanity.get("gini")],
            ["warning", sanity.get("warning")],
        ],
        columns=["metric", "value"],
    )

    top5_df = pd.DataFrame(
        [
            {
                "hazard_id": r.get("hazard_id"),
                "summary": hazard_summary_line(r),
                "recommended_action": r.get("recommended_action"),
                "action_rationale": r.get("action_rationale"),
                "assumption_ids": r.get("assumption_ids"),
                "odd_tags": r.get("odd_tags"),
            }
            for r in rows[:5]
        ]
    )

    cluster_rows = []
    clusters = cluster_hazards(rows)
    for cid, c in sorted(clusters.items(), key=lambda x: len(x[1]["items"]), reverse=True)[:5]:
        rep = c["items"][0] if c["items"] else {}
        cluster_rows.append(
            {
                "cluster_id": cid,
                "cluster_size": len(c["items"]),
                "representative_hazard_id": rep.get("hazard_id"),
            }
        )
    clusters_df = pd.DataFrame(cluster_rows)

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:  # type: ignore
        top_df.to_excel(writer, sheet_name="TopRisks", index=False)
        df.to_excel(writer, sheet_name="Hazards", index=False)
        req_df.to_excel(writer, sheet_name="Requirements", index=False)
        ver_df.to_excel(writer, sheet_name="Verification", index=False)
        tr_df.to_excel(writer, sheet_name="Traceability", index=False)

        meta_df.to_excel(writer, sheet_name="Summary", index=False, startrow=0)
        band_df.to_excel(writer, sheet_name="Summary", index=False, startrow=5)
        cat_df.to_excel(writer, sheet_name="Summary", index=False, startrow=5 + len(band_df) + 3)
        sev_df.to_excel(writer, sheet_name="Summary", index=False, startrow=5 + len(band_df) + len(cat_df) + 6)
        ph_df.to_excel(writer, sheet_name="Summary", index=False, startrow=5 + len(band_df) + len(cat_df) + len(sev_df) + 9)
        trunc_df.to_excel(
            writer,
            sheet_name="Summary",
            index=False,
            startrow=5 + len(band_df) + len(cat_df) + len(sev_df) + len(ph_df) + 12,
        )

        summary_base = 5 + len(band_df) + len(cat_df) + len(sev_df) + len(ph_df) + len(trunc_df) + 16
        top5_df.to_excel(writer, sheet_name="Summary", index=False, startrow=summary_base)
        clusters_df.to_excel(writer, sheet_name="Summary", index=False, startrow=summary_base + len(top5_df) + 4)
        sanity_df.to_excel(
            writer,
            sheet_name="Summary",
            index=False,
            startrow=summary_base + len(top5_df) + len(clusters_df) + 8,
        )

        sys_df.to_excel(writer, sheet_name="SystemInput", index=False)

        wb = writer.book

        def fmt_sheet(name: str, wrap_cols: List[str], width_map: Dict[str, int]) -> None:
            ws = wb[name]
            ws.freeze_panes = "A2"
            header = [cell.value for cell in ws[1]]
            col_idx = {n: i + 1 for i, n in enumerate(header)}

            try:
                from openpyxl.styles import Alignment, Font  # type: ignore

                wrap = Alignment(wrap_text=True, vertical="top")
                top = Alignment(vertical="top")
                bold = Font(bold=True)

                for cell in ws[1]:
                    cell.font = bold

                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    for cell in row:
                        col_name = header[cell.col_idx - 1]
                        cell.alignment = wrap if col_name in wrap_cols else top
            except Exception:
                pass

            for name2, idx in col_idx.items():
                letter = ws.cell(row=1, column=idx).column_letter
                ws.column_dimensions[letter].width = width_map.get(name2, 40)

        fmt_sheet(
            "TopRisks",
            wrap_cols=["trigger_condition", "final_impact", "safety_requirement", "verification_idea", "review_notes", "observables", "mitigations", "propagation_chain", "primary_failure"],
            width_map={
                "hazard_id": 14,
                "risk_rpn": 10,
                "risk_band": 10,
                "category": 14,
                "subtype": 16,
                "task_phase": 18,
                "severity": 10,
                "likelihood": 10,
                "detectability": 12,
                "affected_components": 30,
                "safety_requirement_id": 18,
                "verification_id": 18,
                "why_top": 36,
                "recommended_action": 16,
                "action_rationale": 36,
                "assumption_ids": 24,
                "odd_tags": 40,
                "review_status": 14,
                "reviewer": 14,
                "decision_date": 14,
            },
        )

        fmt_sheet(
            "Hazards",
            wrap_cols=[
                "trigger_condition",
                "primary_failure",
                "propagation_chain",
                "final_impact",
                "observables",
                "measures",
                "mitigations",
                "safety_requirement",
                "verification_idea",
                "review_notes",
                "action_rationale",
                "odd_tags",
                "review_warnings",
            ],
            width_map={
                "hazard_id": 14,
                "risk_rpn": 10,
                "risk_band": 10,
                "category": 14,
                "subtype": 16,
                "task_phase": 18,
                "severity": 10,
                "likelihood": 10,
                "detectability": 12,
                "affected_components": 30,
                "safety_requirement_id": 18,
                "verification_id": 18,
                "assumption_ids": 24,
                "odd_tags": 40,
                "cluster_id": 10,
                "cluster_size": 12,
                "recommended_action": 16,
                "action_rationale": 36,
                "measures": 36,
                "review_warnings": 32,
                "review_status": 14,
                "reviewer": 14,
                "decision_date": 14,
                "flag_req_truncated": 18,
                "flag_verif_truncated": 18,
            },
        )

        fmt_sheet(
            "Requirements",
            wrap_cols=["safety_requirement"],
            width_map={
                "safety_requirement_id": 18,
                "hazard_id": 14,
                "risk_rpn": 10,
                "risk_band": 10,
                "category": 14,
                "task_phase": 18,
                "severity": 10,
                "affected_components": 30,
                "review_status": 14,
                "reviewer": 14,
                "decision_date": 14,
                "flag_req_truncated": 18,
            },
        )

        fmt_sheet(
            "Verification",
            wrap_cols=["verification_idea", "observables", "mitigations"],
            width_map={
                "verification_id": 18,
                "hazard_id": 14,
                "risk_rpn": 10,
                "risk_band": 10,
                "category": 14,
                "task_phase": 18,
                "severity": 10,
                "review_status": 14,
                "reviewer": 14,
                "decision_date": 14,
                "flag_verif_truncated": 18,
            },
        )

        fmt_sheet(
            "Traceability",
            wrap_cols=[],
            width_map={
                "hazard_id": 14,
                "safety_requirement_id": 18,
                "verification_id": 18,
                "risk_rpn": 10,
                "risk_band": 10,
                "severity": 10,
                "task_phase": 18,
                "category": 14,
                "assumption_ids": 24,
                "odd_tags": 40,
                "review_status": 14,
            },
        )


def main() -> None:
    ap = argparse.ArgumentParser(description="Generate XLSX + HTML report from hazards_enriched*.json")
    ap.add_argument("--input", default="outputs/hazards_enriched.json")
    ap.add_argument("--outdir", default="outputs")
    ap.add_argument("--basename", default="report")
    ap.add_argument("--no-xlsx", action="store_true")
    ap.add_argument("--top", type=int, default=20, help="Number of top risks for executive summary")
    args = ap.parse_args()

    in_path = Path(args.input)
    if not in_path.exists():
        raise FileNotFoundError(f"Input not found: {in_path}")

    outdir = Path(args.outdir)
    ensure_dir(outdir)

    doc = load_json(in_path)
    rows = to_rows(doc)
    apply_action_fields(rows)
    clusters = apply_cluster_labels(rows)
    summary = summarize(rows)
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    doc["_generated_at"] = generated_at

    html_path = outdir / f"{args.basename}.html"
    write_html(html_path, doc, rows, summary, args.top, generated_at, clusters)
    print(f"[ok] wrote: {html_path}")

    if not args.no_xlsx:
        xlsx_path = outdir / f"{args.basename}.xlsx"
        try:
            write_xlsx(xlsx_path, rows, summary, doc, args.top)
            print(f"[ok] wrote: {xlsx_path}")
        except Exception as e:
            print(f"[warn] failed to write xlsx: {type(e).__name__}: {e}")
            print("[warn] You can still use the HTML report, or install deps: pip install pandas openpyxl")

    print("=== REPORT QA ===")
    print(f"input_used: {in_path}")
    print(f"total_hazards: {summary['total_hazards']}")
    print(f"risk_band_distribution: {summary['risk_band_distribution']}")
    print(f"truncation_flags: {summary['truncation_flags']}")


if __name__ == "__main__":
    main()
